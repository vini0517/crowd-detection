import os
import cv2
import torch
import numpy as np
from flask import Flask, request, jsonify, render_template
from ultralytics import YOLO
from transformers import DPTForDepthEstimation, DPTImageProcessor
from PIL import Image
from openpyxl import Workbook, load_workbook
from datetime import datetime
import traceback  # For detailed error logging

# Initialize Flask app
app = Flask(__name__)

# Define upload and processed folders
UPLOAD_FOLDER = "static/uploads"
PROCESSED_FOLDER = "static/processed"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PROCESSED_FOLDER, exist_ok=True)

# Maximum allowed file size (10 MB)
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB in bytes

# Maximum allowed image dimensions (width x height)
MAX_IMAGE_WIDTH = 1920  # Maximum width
MAX_IMAGE_HEIGHT = 1080  # Maximum height

# Load YOLO model for face detection
face_model = YOLO("crowddetection.pt")  # Replace with your YOLO model path

# Load Depth Estimation model
depth_model = DPTForDepthEstimation.from_pretrained("Intel/dpt-large").eval()
feature_extractor = DPTImageProcessor.from_pretrained("Intel/dpt-large")

# Function to save analysis data to Excel
def save_to_excel(data, image_name, filename="analysis_data.xlsx"):
    # Check if the file already exists
    if os.path.exists(filename):
        # Load the existing workbook
        wb = load_workbook(filename)
        ws = wb.active

        # Check if the image_name already exists in the file
        for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header row
            if row[1] == image_name:  # row[1] is the "Image Name" column
                print(f"Data for {image_name} already exists. Skipping...")
                return
    else:
        # Create a new workbook if it doesn't exist
        wb = Workbook()
        ws = wb.active
        # Add headers
        ws.append(["Timestamp", "Image Name", "Total Crowd", "Risk Level", "Zone Counts"])

    # Append new data
    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),  # Timestamp
        image_name,  # Image Name
        data.get("total_count", 0),  # Total Crowd
        data.get("risk_level", "-"),  # Risk Level
        str(data.get("zone_counts", "-"))  # Zone Counts
    ])

    # Save the workbook
    wb.save(filename)

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/upload", methods=["POST"])
def upload():
    if "image" not in request.files:
        return jsonify({"error": "No image file provided"}), 400

    image_file = request.files["image"]
    filename = image_file.filename
    image_path = os.path.join(UPLOAD_FOLDER, filename)

    # Check file size
    image_file.seek(0, os.SEEK_END)  # Move cursor to end of file
    file_size = image_file.tell()  # Get file size
    image_file.seek(0)  # Reset cursor to beginning of file

    if file_size > MAX_FILE_SIZE:
        return jsonify({"error": "File size exceeds the maximum allowed limit (10 MB)"}), 400

    # Save the file
    image_file.save(image_path)

    try:
        # Load image using PIL
        image = Image.open(image_path).convert("RGB")
        print(f"Image format: {image.format}")  # Debugging

        # Check image dimensions
        width, height = image.size
        print(f"Image dimensions: {width}x{height}")  # Debugging
        if width > MAX_IMAGE_WIDTH or height > MAX_IMAGE_HEIGHT:
            # Resize the image to fit within the maximum dimensions
            image.thumbnail((MAX_IMAGE_WIDTH, MAX_IMAGE_HEIGHT), Image.Resampling.LANCZOS)
            image.save(image_path)  # Save the resized image

        # Load the resized image using OpenCV
        img_cv = cv2.imread(image_path)
        if img_cv is None:
            return jsonify({"error": "Invalid image file"}), 400

        # Step 1: Detect Faces using YOLO
        results = face_model(image, conf=0.1)
        detections = results[0].boxes.xyxy.cpu().numpy()  # Get bounding boxes

        total_faces = len(detections)
        print(f"Total faces detected: {total_faces}")  # Debugging
        if total_faces == 0:
            return jsonify({"total_faces": 0, "message": "No faces detected"})

        # Draw Bounding Boxes for Face Detections (Green)
        for box in detections:
            xmin, ymin, xmax, ymax = map(int, box)
            cv2.rectangle(img_cv, (xmin, ymin), (xmax, ymax), (0, 255, 0), 3)  # Green box
            cv2.putText(img_cv, "Face", (xmin, ymin - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)

        # Step 2: Generate Depth Map
        inputs = feature_extractor(images=image, return_tensors="pt")
        with torch.no_grad():
            depth = depth_model(**inputs).predicted_depth
        depth_map = depth.squeeze().numpy()

        # Resize depth map to match the original image dimensions
        depth_map = cv2.resize(depth_map, (image.width, image.height), interpolation=cv2.INTER_LINEAR)

        # Normalize depth map to [0, 255]
        depth_map = (depth_map - depth_map.min()) / (depth_map.max() - depth_map.min()) * 255
        depth_map = depth_map.astype(np.uint8)

        # Step 3: Generate Density Map using Depth Information
        density_map = np.zeros((img_cv.shape[0], img_cv.shape[1]), dtype=np.float32)
        for box in detections:
            xmin, ymin, xmax, ymax = map(int, box)
            center_x = (xmin + xmax) // 2
            center_y = (ymin + ymax) // 2

            # Ensure coordinates are within bounds
            if 0 <= center_y < depth_map.shape[0] and 0 <= center_x < depth_map.shape[1]:
                depth_value = depth_map[center_y, center_x]
                # Closer objects have higher weight (inverse of depth value)
                weight = 255.0 - depth_value  # Normalize to [0, 255]
                cv2.circle(density_map, (center_x, center_y), 15, weight, thickness=-1)

        # Apply Gaussian blur to smooth the density map
        density_map = cv2.GaussianBlur(density_map, (21, 21), 0)

        # Normalize density map to [0, 255]
        if density_map.max() != 0:
            density_map = (density_map / density_map.max()) * 255
        density_map = density_map.astype(np.uint8)

        # Step 4: Find Very High-Density Regions
        # Use a higher threshold to identify very high-density regions
        _, binary_density_map = cv2.threshold(density_map, 220, 255, cv2.THRESH_BINARY)  # Adjusted threshold
        binary_density_map = binary_density_map.astype(np.uint8)

        # Find contours of very high-density regions
        contours, _ = cv2.findContours(binary_density_map, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        # Filter out small contours (e.g., areas with fewer than 4 people)
        min_area = 5000  # Adjust this value based on your image resolution
        valid_contours = [cnt for cnt in contours if cv2.contourArea(cnt) > min_area]

        # Draw red boxes around valid high-density regions
        is_high_density = False
        for contour in valid_contours:
            x, y, w, h = cv2.boundingRect(contour)
            cv2.rectangle(img_cv, (x, y), (x + w, y + h), (0, 0, 255), 4)  # Red box for very high-density region
            cv2.putText(img_cv, "High Density", (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)
            is_high_density = True  # Set flag if red boxes are drawn

        # Calculate the zone count
        zone_counts = len(valid_contours)

        # Step 5: Check Overcrowding and High-Density Conditions
        max_threshold = 200  # Set your maximum threshold for overcrowding
        is_overcrowded = total_faces > max_threshold

        # Step 6: Report Overcrowded Areas to Security
        if is_high_density:
            print(f"Very high-density areas detected in {filename}. Notifying security...")
            # Here you can add code to notify security (e.g., send an email, log to a file, etc.)

        # Step 7: Save Processed Image
        processed_filename = "processed_" + filename
        processed_image_path = os.path.join(PROCESSED_FOLDER, processed_filename)
        cv2.imwrite(processed_image_path, img_cv)

        # Save Density Map for visualization (No bounding boxes)
        density_map_colored = cv2.applyColorMap(density_map, cv2.COLORMAP_JET)
        density_filename = "density_" + filename
        density_image_path = os.path.join(PROCESSED_FOLDER, density_filename)
        cv2.imwrite(density_image_path, density_map_colored)

        # Save Depth Map for visualization (No bounding boxes)
        depth_map_colored = cv2.applyColorMap(depth_map, cv2.COLORMAP_JET)
        depth_filename = "depth_" + filename
        depth_image_path = os.path.join(PROCESSED_FOLDER, depth_filename)
        cv2.imwrite(depth_image_path, depth_map_colored)

        # Save Binary Density Map for visualization
        binary_density_map_colored = cv2.applyColorMap(binary_density_map, cv2.COLORMAP_JET)
        binary_density_filename = "binary_density_" + filename
        binary_density_image_path = os.path.join(PROCESSED_FOLDER, binary_density_filename)
        cv2.imwrite(binary_density_image_path, binary_density_map_colored)

        # Prepare analysis data
        analysis_data = {
            "total_count": total_faces,
            "is_overcrowded": is_overcrowded,  # Flag for overcrowding
            "is_high_density": is_high_density,  # Flag for very high-density regions
            "risk_level": "High" if is_overcrowded or is_high_density else "Low",  # Risk level
            "zone_counts": len(contours)  # Number of very high-density areas
        }

        # Save analysis data to Excel
        save_to_excel(analysis_data, filename)  # Pass the image name

        # Return JSON response
        return jsonify({
            "total_count": total_faces,
            "is_overcrowded": is_overcrowded,
            "is_high_density": is_high_density,
            "processed_image": processed_filename,
            "density_map": density_filename,
            "depth_map": depth_filename,
            "binary_density_map": binary_density_filename
        })

    except Exception as e:
        # Log the full error traceback
        print(f"Error: {str(e)}")
        print(traceback.format_exc())  # Print the full traceback
        return jsonify({"error": "An error occurred! Please check the logs for details."}), 500

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))  # Render dynamically assigns port
    app.run(host='0.0.0.0', port=port)
